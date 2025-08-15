# %%
import pandas as pd
import os
import glob
import numpy as np
from scipy.stats import pearsonr
from sklearn.metrics import mean_squared_error
from openpyxl import load_workbook

# Definisikan nama AWS
nama_aws = "AWS_Maritim_Cilacap"

# Direktori input dan output
input_dir = r"D:\anto\proposal\modulskripsi\verifikasi\dataanginmodel\{}".format(nama_aws)
output_dir = os.path.join(input_dir, "gabungan_korelasi_dan_rmse_terdekat")
os.makedirs(output_dir, exist_ok=True)

# Fungsi untuk membersihkan data AWS
def clean_aws_data(df):
    print("Sebelum pembersihan AWS:", len(df), "baris")
    df['Tanggal'] = pd.to_datetime(df['Tanggal'], errors='coerce').dt.tz_localize(None)
    df = df[~((df['ws_avg_flag'] == '9') | (df['wd_avg_flag'] == '9') | (df['ws_avg'] == '9') | (df['wd_avg'] == '9'))]
    df = df.dropna(subset=['Tanggal', 'ws_avg', 'wd_avg'])
    print("Setelah pembersihan AWS:", len(df), "baris")
    return df

# Fungsi untuk membersihkan data Titik Terdekat
def clean_wind_data(df):
    print("Sebelum pembersihan Titik Terdekat:", len(df), "baris")
    df['time'] = pd.to_datetime(df['time'], errors='coerce').dt.tz_localize(None)
    df = df[~((df['wind_speed_ms'] == 9) | (df['wind_dir_deg'] == '9'))]
    df = df.dropna(subset=['time', 'wind_speed_ms', 'wind_dir_deg'])
    print("Setelah pembersihan Titik Terdekat:", len(df), "baris")
    return df

# Fungsi untuk menghitung korelasi dan RMSE
def calculate_metrics(ws_avg, titikterdekat_ws):
    valid_data = pd.DataFrame({'ws_avg': ws_avg, 'titikterdekat_ws': titikterdekat_ws}).dropna()
    if len(valid_data) > 1:
        corr, _ = pearsonr(valid_data['ws_avg'], valid_data['titikterdekat_ws'])
        rmse = np.sqrt(mean_squared_error(valid_data['ws_avg'], valid_data['titikterdekat_ws']))
    else:
        corr = np.nan
        rmse = np.nan
    return corr, rmse

# Fungsi untuk menemukan sheet AWS yang sesuai
def find_aws_sheet(aws_file, period):
    wb = load_workbook(aws_file, read_only=True)
    sheet_names = wb.sheetnames
    # Cari sheet dengan nama persis atau awalan sesuai periode (YYYYMM)
    for sheet_name in sheet_names:
        if sheet_name.startswith(period[:6]) or sheet_name == period:
            return sheet_name
    return None

# Mendapatkan semua file wind_data_*.xlsx
wind_files = glob.glob(os.path.join(input_dir, "wind_data_*.xlsx"))

for wind_file in wind_files:
    # Mendapatkan nama periode dari nama file
    period = os.path.basename(wind_file).replace("wind_data_", "").replace(".xlsx", "")
    print(f"\nMemproses file: {wind_file} (periode: {period})")
    
    # Membaca sheet Titik Terdekat
    try:
        titikterdekat = pd.read_excel(wind_file, sheet_name="Titik Terdekat")
        titikterdekat = clean_wind_data(titikterdekat)
        print(f"Data Titik Terdekat untuk {period}: {len(titikterdekat)} baris")
    except Exception as e:
        print(f"Gagal membaca sheet Titik Terdekat dari {wind_file}: {e}")
        titikterdekat = pd.DataFrame()
    
    # Membaca data AWS dari sheet yang sesuai
    aws_file = os.path.join(input_dir, f"{nama_aws}_merged.xlsx")
    sheet_name = find_aws_sheet(aws_file, period)
    if sheet_name:
        try:
            aws_data = pd.read_excel(aws_file, sheet_name=sheet_name)
            aws_data = clean_aws_data(aws_data)
            print(f"Data AWS dari sheet {sheet_name}: {len(aws_data)} baris")
            print(f"Rentang waktu AWS: {aws_data['Tanggal'].min()} sampai {aws_data['Tanggal'].max()}")
        except Exception as e:
            print(f"Gagal membaca sheet {sheet_name} dari {aws_file}: {e}")
            aws_data = pd.DataFrame()
    else:
        print(f"Tidak ditemukan sheet untuk periode {period} di {aws_file}")
        aws_data = pd.DataFrame()
    
    # Filter data AWS untuk periode yang sesuai
    if not titikterdekat.empty:
        start_date = titikterdekat['time'].min()
        end_date = titikterdekat['time'].max()
        print(f"Rentang waktu Titik Terdekat: {start_date} sampai {end_date}")
        if not aws_data.empty:
            aws_period = aws_data[(aws_data['Tanggal'] >= start_date) & (aws_data['Tanggal'] <= end_date)]
            print(f"Data AWS untuk periode {period} setelah filter: {len(aws_period)} baris")
        else:
            aws_period = pd.DataFrame(columns=['Nama Stasiun', 'Latitude', 'Longitude', 'Tanggal', 'ws_avg', 'wd_avg'])
            print(f"Tidak ada data AWS untuk periode {period} setelah filter")
    else:
        aws_period = pd.DataFrame(columns=['Nama Stasiun', 'Latitude', 'Longitude', 'Tanggal', 'ws_avg', 'wd_avg'])
        print(f"Tidak ada data Titik Terdekat untuk {period}, AWS diisi kosong")
    
    # Menggabungkan data untuk sheet Gabungan
    merged_data = aws_period[['Nama Stasiun', 'Latitude', 'Longitude', 'Tanggal', 'ws_avg', 'wd_avg']].copy()
    if not titikterdekat.empty:
        titikterdekat_subset = titikterdekat[['time', 'wind_speed_ms', 'wind_dir_deg']].rename(columns={
            'wind_speed_ms': 'titikterdekat_wind_speed_ms',
            'wind_dir_deg': 'titikterdekat_wind_dir_deg'
        })
        merged_data = merged_data.merge(
            titikterdekat_subset,
            how='left',
            left_on='Tanggal',
            right_on='time'
        )
        merged_data = merged_data.drop(columns=['time'], errors='ignore')
        
        # Menghitung korelasi dan RMSE
        corr, rmse = calculate_metrics(merged_data['ws_avg'], merged_data['titikterdekat_wind_speed_ms'])
        merged_data[''] = np.nan  # Kolom kosong
        merged_data['Correlation_ws'] = np.nan
        merged_data['RMSE_ws'] = np.nan
        if len(merged_data) > 0:
            merged_data.iloc[0, merged_data.columns.get_loc('Correlation_ws')] = corr
            merged_data.iloc[0, merged_data.columns.get_loc('RMSE_ws')] = rmse
    else:
        merged_data['titikterdekat_wind_speed_ms'] = np.nan
        merged_data['titikterdekat_wind_dir_deg'] = np.nan
        merged_data[''] = np.nan
        merged_data['Correlation_ws'] = np.nan
        merged_data['RMSE_ws'] = np.nan
    
    # Menyimpan ke file Excel dengan nama sheet ketiga YYYYMM
    output_file = os.path.join(output_dir, f"combined_{period}.xlsx")
    sheet_name = period[:6]  # Gunakan YYYYMM, misalnya 202106
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        titikterdekat.to_excel(writer, sheet_name='Titik Terdekat', index=False)
        aws_period.to_excel(writer, sheet_name='AWS', index=False)
        merged_data.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"File Excel untuk periode {period} telah disimpan di: {output_file}")

# Rangkuman correlation dan rmse
# Direktori input dan output
input_dir = r"D:\anto\proposal\modulskripsi\verifikasi\dataanginmodel\{}\gabungan_korelasi_dan_rmse_terdekat".format(nama_aws)
output_file = os.path.join(input_dir, f"correlation_rmse_summary_terdekat_{nama_aws}.xlsx")

# Inisialisasi daftar untuk menyimpan hasil
results = []

# Mendapatkan semua file combined_*.xlsx
combined_files = glob.glob(os.path.join(input_dir, "combined_*.xlsx"))

for file in combined_files:
    # Mendapatkan nama periode dari nama file
    period = os.path.basename(file).replace("combined_", "").replace(".xlsx", "")[:6]  # Gunakan YYYYMM
    print(f"Memproses file: {file} (periode: {period})")
    
    # Membaca sheet ketiga (dinamakan sesuai periode YYYYMM)
    try:
        df = pd.read_excel(file, sheet_name=period)
        # Mengambil Correlation_ws dan RMSE_ws dari baris pertama
        correlation = df['Correlation_ws'].iloc[0] if 'Correlation_ws' in df.columns and not pd.isna(df['Correlation_ws'].iloc[0]) else None
        rmse = df['RMSE_ws'].iloc[0] if 'RMSE_ws' in df.columns and not pd.isna(df['RMSE_ws'].iloc[0]) else None
        results.append({
            'Periode': period,
            'Correlation_ws': correlation,
            'RMSE_ws': rmse
        })
        print(f"Berhasil membaca sheet {period}: Correlation_ws={correlation}, RMSE_ws={rmse}")
    except Exception as e:
        print(f"Gagal membaca sheet {period} dari file {file}: {e}")
        results.append({
            'Periode': period,
            'Correlation_ws': None,
            'RMSE_ws': None
        })

# Menggabungkan hasil ke DataFrame
results_df = pd.DataFrame(results)
results_df = results_df.sort_values('Periode').reset_index(drop=True)

# Menyimpan ke file Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    results_df.to_excel(writer, sheet_name='Summary', index=False)

print(f"File Excel telah disimpan di: {output_file}")


